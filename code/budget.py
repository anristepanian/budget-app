import pandas as pd
import openpyxl as xsl
from openpyxl.utils import get_column_letter
import os
from datetime import date as dt
from datetime import timedelta
from db import DB


class Budget:
    @staticmethod
    def data_drop(db, date: str, income: bool = False, expense: int = False):
        if not income and not expense:
            print(income, expense)
            return print("What was the purpose of this??")
        wb = xsl.load_workbook('budget.xlsx')
        ws = wb['Data']
        date_dt = dt.fromisoformat(date)
        if income:
            ws[f"C{str(date_dt.day + 3)}"].value = 0
            DB.delete_income(db, date)
            print("Your income is successfully deleted!")
        if expense:
            ws[f"D{str(date_dt.day + 3)}"].value = 0
            DB.delete_expense(db, date)
            print("Your expense is successfully deleted!")
        wb.save('budget.xlsx')
        wb.close()

    @staticmethod
    def data_entry(db, date: str, income: float = None, income_category: str = None, expense: float = None,
                   expense_category: str = None):
        if not income and not expense:
            return print("What was the purpose of this??")
        wb = xsl.load_workbook('budget.xlsx', data_only=True)
        ws = wb['Data']
        date_dt = dt.fromisoformat(date)
        if income != .0:
            ws[f"C{str(date_dt.day + 3)}"].value = float(ws[f"C{str(date_dt.day + 3)}"].value) + float(income)
            print("Your income is successfully entered!")
        if expense != .0:
            ws[f"D{str(date_dt.day + 3)}"].value = float(ws[f"D{str(date_dt.day + 3)}"].value) + float(expense)
            print("Your expense is successfully entered!")
        for i in range(4, 36):
            ws[f"E{i}"].value = f'=C{i}-D{i}'
        ws[f"C35"].value = '=SUM(C4:C34)'
        ws[f"D35"].value = '=SUM(D4:D34)'
        wb.save('budget.xlsx')
        wb.close()
        DB.add_action(db, date, income, income_category, expense, expense_category)
        if (date_dt + timedelta(days=1)).month == date_dt.month + 1:
            return print("Today is the last day of the month!\nWould you like to start a new month?")

    @staticmethod
    def new_month(db, cur_year, cur_month):
        cur_month -= 1
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']

        wb = xsl.load_workbook(filename='budget.xlsx', data_only=True)
        sheet_names = wb.get_sheet_names()
        name = sheet_names[0]
        sheet_ranges = wb[name]
        df = pd.DataFrame(sheet_ranges.values)

        folder_path = f'{cur_year}'
        os.makedirs(folder_path, exist_ok=True)
        csv_file = f'{folder_path}/{months[cur_month]}_{cur_year}.csv'
        df.to_csv(csv_file, index=False, header=False)

        wb = xsl.load_workbook('budget.xlsx', data_only=True)
        ws = wb['Data']
        total = ws["E35"].value
        if not total:
            total = 0
        values = [.0] * 12
        values[cur_month] = total
        DB.add_month(db, cur_year, values)

        if cur_month == 11:
            DB.add_total(db, cur_year)
        wb.close()
        wb = xsl.load_workbook('budget.xlsx')
        ws = wb['Data']

        for col in range(3, 5):
            for row in range(4, 35):
                char = get_column_letter(col)
                ws[char + str(row)].value = 0

        wb.save('budget.xlsx')
        wb.close()

        print("You have successfully started a new month!")
